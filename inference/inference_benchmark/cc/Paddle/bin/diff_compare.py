import argparse
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter


def parse_args():
    """
    parse input args
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--new_excel", type=str, default="./benchmark_model_excel.xlsx",
                        help="new excel")
    parser.add_argument("--last_excel", type=str, default="./benchmark_model_excel_last.xlsx",
                        help="last time excel")
    parser.add_argument("--output_name", type=str, default="benchmark_model_diff.xlsx",
                        help="output excel file name")
    return parser.parse_args()


def color_mark(cells, threshold, reversal=False):
    """
    mark cell color
    :param cells:
    :return: None
    """
    try:
        for cell in cells:
            if cell.value is None:
                continue
            if cell.value < -threshold:
                cell.font = Font(color=Color(rgb="1aad19")) if reversal else Font(color=Color(index=2))
            elif cell.value > threshold:
                cell.font = Font(color=Color(index=2)) if reversal else Font(color=Color(rgb="1aad19"))
    except Exception as e:
        pass


def set_style(diff_excel):
    """
    set excel style
    """
    workbook = load_workbook(diff_excel)
    sheet1 = workbook.active
    cells = sheet1["A:S"]
    # center
    aligncenter = Alignment(horizontal='center', vertical='center')
    for i in cells:
        for j in i:
            j.alignment = aligncenter
    # color mark
    diff_cells_qps = sheet1["O"][1:]
    color_mark(diff_cells_qps, 5, reversal=False)
    diff_cells_cpu_rss_percent = sheet1["P"][1:]
    color_mark(diff_cells_cpu_rss_percent, 10, reversal=True)
    diff_cells_cpu_rss_value = sheet1["Q"][1:]
    color_mark(diff_cells_cpu_rss_value, 100, reversal=True)
    diff_cells_gpu_rss_percent = sheet1["R"][1:]
    color_mark(diff_cells_gpu_rss_percent, 10, reversal=True)
    diff_cells_gpu_rss_value = sheet1["S"][1:]
    color_mark(diff_cells_gpu_rss_value, 100, reversal=True)
    try:
        for cell in diff_cells_qps:
            if cell.value is None:
                continue
            if cell.value < -5:
                cell.font = Font(color=Color(index=2))
            elif cell.value > 5:
                cell.font = Font(color=Color(rgb="1aad19"))
    except Exception as e:
        pass
    # edit width
    for index, i in enumerate(cells):
        column_width = 0
        for j in i:
            if j.value:
                if len(str(j.value)) > column_width:
                    column_width = len(str(j.value))
        sheet1.column_dimensions[get_column_letter(index + 1)].width = column_width + 2

    workbook.save(diff_excel)


if __name__ == '__main__':
    args = parse_args()
    new_df = pd.read_excel(args.new_excel).drop(["Unnamed: 0", "model_type", "num_samples", "ir_optim", "enable_mkldnn",
                                                 "cpu_math_library_num_threads", "cpu_vms(MB)",
                                                 "cpu_shared(MB)", "cpu_dirty(MB)", "cpu_usage(%)",
                                                 "gpu_utilization_rate(%)", "gpu_mem_utilization_rate(%)"], axis=1)
    new_qps_df = new_df["QPS"]

    last_df = pd.read_excel(args.last_excel).drop(["Unnamed: 0", "model_type", "num_samples", "ir_optim", "enable_mkldnn",
                                                    "cpu_math_library_num_threads", "cpu_vms(MB)",
                                                    "cpu_shared(MB)", "cpu_dirty(MB)", "cpu_usage(%)",
                                                    "gpu_utilization_rate(%)", "gpu_mem_utilization_rate(%)"], axis=1)
    last_qps_df = last_df["QPS"]
    merge_df = pd.merge(new_df, last_df, how="inner",
                        on=["model_name", "batch_size", "enable_tensorrt", "trt_precision"],
                        suffixes=("", "_last")).drop(["device_last"], axis=1)
    merge_df["QPS_diff(%)"] = merge_df[["QPS", "QPS_last"]].apply(lambda x: (x["QPS"] - x["QPS_last"]) / x["QPS_last"] * 100, axis=1)
    merge_df["cpu_rss_diff(%)"] = merge_df[["cpu_rss(MB)", "cpu_rss(MB)_last"]].apply(lambda x: (x["cpu_rss(MB)_last"] - x["cpu_rss(MB)"]) / x["cpu_rss(MB)"] * 100, axis=1)
    merge_df["cpu_rss_diff(MB)"] = merge_df[["cpu_rss(MB)", "cpu_rss(MB)_last"]].apply(lambda x: x["cpu_rss(MB)_last"] - x["cpu_rss(MB)"], axis=1)
    merge_df["gpu_mem_diff(%)"] = merge_df[["gpu_mem(MB)", "gpu_mem(MB)_last"]].apply(lambda x: (x["gpu_mem(MB)_last"] - x["gpu_mem(MB)"]) / x["gpu_mem(MB)"] * 100, axis=1)
    merge_df["gpu_mem_diff(MB)"] = merge_df[["gpu_mem(MB)", "gpu_mem(MB)_last"]].apply(lambda x: x["gpu_mem(MB)_last"] - x["gpu_mem(MB)"], axis=1)
    print(merge_df)
    merge_df.sort_values(by=["model_name", "batch_size", "trt_precision"], inplace=True)
    merge_df = merge_df.dropna(subset=["model_name"])
    merge_df.to_excel(args.output_name)
    # postprocess
    set_style(args.output_name)

